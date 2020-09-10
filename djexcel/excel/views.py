import json

from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from openpyxl import Workbook
from excel.models import *
from threading import Timer
# Create your views here.

def exceltest(request):
    return render(request, 'testcel.html',{})

@csrf_exempt
def salary_update_excel(request):
    if request.method == 'POST':

        # 파일 저장
        # file = request.FILES['file_excel']
        # fs = FileSystemStorage()
        # filename = fs.save(file.name, file)
        # uploaded_file_url = fs.url(filename)
        # print(uploaded_file_url)

        file = request.FILES['file_excel']
        # data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
        load_wb = load_workbook(file, data_only=True)
        # 시트 이름으로 불러오기
        load_ws = load_wb['Sheet1']
        # 셀 주소로 값 출력
        # print(load_ws['A1'].value)

        # 일단 리스트에 담기
        all_values = []
        for row in load_ws.rows:
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)

        cnt = 0
        for idx, val in enumerate(all_values):
            if idx == 0:
                # 엑셀 형식 체크 (첫번째의 제목 row)
                if val[0] != '이름' or val[1] != '수량' or val[2] != '가격' or val[3] != '금액':
                    context = {'state': False, 'rtnmsg': '엑셀 항목이 올바르지 않습니다.'}
                    return HttpResponse(json.dumps(context), content_type="application/json")
            else:
                try:
                    memData = Member.objects.create(name=val[0], count=val[1], price=val[2], total=val[3])
                    memData.save()
                    cnt += 1
                except:
                    print('중복제거')
        context = {'state': True, 'rtnmsg': '{0}건의 엑셀 데이터가 반영 되었습니다.'.format(cnt)}
        return HttpResponse(json.dumps(context), content_type="application/json")


def excelcreate(request):
    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws['A1'] = '이름'
    write_ws['A2'] = '수량'
    write_ws['A3'] = '가격'
    write_ws['A4'] = '금액'
    context = {'state':False, 'rtnmsg':'test'}
    return HttpResponse(json.dumps(context), content_type="application/json")